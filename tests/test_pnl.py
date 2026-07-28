"""Tests for the paper-trade P&L summary."""

from __future__ import annotations

from datetime import UTC, datetime, timedelta

from arb.analysis.pnl import export_excel, summarize
from arb.core.matching import CanonicalPair, OutcomeMapping
from arb.core.pricing import ArbLeg, ArbOpportunity
from arb.infra.db import Database
from arb.live.simulator import PaperLedger
from arb.providers.models import Outcome, Venue


def _pair():
    return CanonicalPair(
        "fed-cut", "Fed cut",
        {"kalshi": OutcomeMapping("kalshi", "KX-1", Outcome.YES),
         "polymarket": OutcomeMapping("polymarket", "0xabc", Outcome.YES)},
    )


def _opp():
    # 60 pairs: cost 24+33=57, fee 1.0 -> net = 60 - 57 - 1 = 2.0
    return ArbOpportunity(
        contracts=60, gross_profit=3.0, net_profit=2.0, roi=2.0 / 57, net_edge_per_contract=2.0 / 60,
        legs=(
            ArbLeg(Venue.KALSHI, Outcome.YES, 60, 0.40, 24.0, 1.0),
            ArbLeg(Venue.POLYMARKET, Outcome.NO, 60, 0.55, 33.0, 0.0),
        ),
    )


def test_summary_empty(tmp_path):
    db = Database(tmp_path / "p.sqlite"); db.init_schema()
    assert summarize(db).is_empty


def test_summary_reconstructs_pnl_from_fills(tmp_path):
    db = Database(tmp_path / "p.sqlite"); db.init_schema()
    ledger = PaperLedger(db=db)
    t0 = datetime.now(UTC)
    ledger.record(_pair(), _opp(), t0)
    ledger.record(_pair(), _opp(), t0 + timedelta(seconds=1))  # distinct ts -> 2nd trade

    summ = summarize(db)
    assert summ.trades == 2
    assert summ.net_profit == 2.0 * 2  # recomputed from fills, not the stored note
    assert summ.contracts == 120  # 60 per trade
    assert summ.fees == 2.0  # 1.0 kalshi fee per trade
    assert summ.gross_cost == 57.0 * 2

    # per-pair rollup
    assert set(summ.by_pair["canonical_id"]) == {"fed-cut"}
    assert float(summ.by_pair.iloc[0]["net_profit"]) == 4.0

    # per-venue fees: kalshi carries the fee, polymarket zero
    by_venue = {r["venue"]: r for _, r in summ.by_venue.iterrows()}
    assert by_venue["kalshi"]["fees"] == 2.0
    assert by_venue["polymarket"]["fees"] == 0.0


def test_export_excel(tmp_path):
    import openpyxl

    db = Database(tmp_path / "p.sqlite"); db.init_schema()
    ledger = PaperLedger(db=db)
    t0 = datetime.now(UTC)
    ledger.record(_pair(), _opp(), t0)
    ledger.record(_pair(), _opp(), t0 + timedelta(seconds=1))

    out = export_excel(db, tmp_path / "pnl.xlsx")
    assert out.exists()
    wb = openpyxl.load_workbook(out)
    assert {"summary", "by_pair", "by_venue", "trades"} <= set(wb.sheetnames)
    # summary sheet: header + one data row with the total net.
    ws = wb["summary"]
    header = [c.value for c in ws[1]]
    row = [c.value for c in ws[2]]
    assert row[header.index("net_profit")] == 4.0
    assert row[header.index("trades")] == 2
    # trades sheet has both trades.
    assert wb["trades"].max_row == 3  # header + 2 trades


def test_export_excel_empty(tmp_path):
    db = Database(tmp_path / "e.sqlite"); db.init_schema()
    out = export_excel(db, tmp_path / "empty.xlsx")
    assert out.exists()  # writes a valid (zeroed) workbook even with no trades
